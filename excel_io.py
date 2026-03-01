from openpyxl import load_workbook
from typing import List
from dataclasses import dataclass

# ─────────────────────────────────────────────────────────────────────────────
# 2. DATA STRUCTURES
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class TenderDocument:
    tender_id: str
    budget: float
    delivery_months: int
    min_annual_turnover: float
    required_certifications: List[str]
    min_years_experience: int
    min_public_liability: float
    min_past_contract_value: float
    min_employees: int
    evaluation_criteria: str


@dataclass
class Bid:
    bid_id: str
    company_name: str
    bid_price: float
    proposed_delivery_months: int
    annual_turnover: float
    certifications: List[str]
    years_experience: int
    public_liability_cover: float
    largest_past_contract: float
    employee_count: int
    technical_approach: str
    added_value: str
    notes: str = ""


@dataclass
class ReasoningStep:
    label: str
    finding: str
    conclusion: str


@dataclass
class BidEvaluation:
    bid: Bid
    reasoning_steps: List[ReasoningStep]
    qualification_passed: bool
    scores: dict
    total_score: int
    recommendation_score: int
    confidence: str
    recommendation: str
    recommendation_reason: str
    summary: str



def load_tender(path: str) -> TenderDocument:
    wb = load_workbook(path, data_only=True)
    sheet = wb["Tender"]

    row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]

    if not row or not row[0]:
        raise ValueError("Tender sheet is empty or misformatted.")

    return TenderDocument(
        tender_id=row[0],
        budget=float(row[1]),
        delivery_months=int(row[2]),
        evaluation_criteria=row[3],
        min_annual_turnover=float(row[4]),
        required_certifications=[c.strip() for c in (row[5] or "").split(",") if c.strip()],
        min_years_experience=int(row[6]),
        min_public_liability=float(row[7]),
        min_past_contract_value=float(row[8]),
        min_employees=int(row[9] or 0),
    )


def load_bids(path: str) -> List[Bid]:
    wb = load_workbook(path, data_only=True)
    sheet = wb["Bids"]

    bids = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        bids.append(Bid(
            bid_id=row[0],
            company_name=row[1],
            bid_price=float(row[2]),
            proposed_delivery_months=int(row[3]),
            annual_turnover=float(row[4]),
            certifications=[c.strip() for c in (row[5] or "").split(",") if c.strip()],
            years_experience=int(row[6]),
            public_liability_cover=float(row[7]),
            largest_past_contract=float(row[8]),
            employee_count=int(row[9]),
            technical_approach=row[10] or "",
            added_value=row[11] or "",
            notes=row[12] or "",
        ))

    return bids


def write_results(path: str, evaluations: List[BidEvaluation]):
    wb = load_workbook(path)

    if "Results" in wb.sheetnames:
        del wb["Results"]

    sheet = wb.create_sheet("Results")

    headers = [
        "Rank",
        "Bid ID",
        "Company",
        "Qualified",
        "Price",
        "Delivery",
        "Technical",
        "Compliance",
        "Added Value",
        "Total",
        "Recommendation Score",
        "Verdict",
        "Confidence"
    ]

    sheet.append(headers)

    # Rank: Qualified first, then by recommendation score desc
    sorted_evals = sorted(
        evaluations,
        key=lambda e: (0 if e.qualification_passed else 1, -e.recommendation_score)
    )

    for rank, ev in enumerate(sorted_evals, 1):
        sheet.append([
            rank,
            ev.bid.bid_id,
            ev.bid.company_name,
            "YES" if ev.qualification_passed else "NO",
            ev.scores.get("price", 0),
            ev.scores.get("delivery", 0),
            ev.scores.get("technical", 0),
            ev.scores.get("compliance", 0),
            ev.scores.get("added_value", 0),
            ev.total_score,
            ev.recommendation_score,
            ev.recommendation,
            ev.confidence
        ])

    wb.save(path)